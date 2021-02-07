# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook
import urllib

def updateExcelFile():
    load_wb = load_workbook('../MapleGuildSuro/party.xlsx', data_only=True)
    load_ws = load_wb['GuildMemberList']

    for i in range(3, load_ws.max_row+1):
        html = urlopen("https://maple.gg/u/" + urllib.parse.quote(load_ws.cell(row=i, column=1).value))
        bsObject = BeautifulSoup(html, "html.parser")

        load_ws.cell(i, 2, bsObject.find('ul', class_='user-summary-list').find_all('li', class_='user-summary-item')[0].get_text().replace("Lv.",""))
        load_ws.cell(i, 3, bsObject.find('ul', class_='user-summary-list').find_all('li', class_='user-summary-item')[1].get_text())
        load_ws.cell(i, 4, bsObject.find('h1', class_='user-summary-floor font-weight-bold').get_text().replace(" ","").replace("\n","").replace("층",""))
        load_ws.cell(i, 5, bsObject.find('small', class_='user-summary-duration').get_text())

    load_wb.save('../MapleGuildSuro/party.xlsx')

if __name__ == '__main__':
    updateExcelFile()